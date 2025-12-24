import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.metrics import mean_squared_error, mean_absolute_error
from fpdf import FPDF
import io
from openpyxl import Workbook
from prophet import Prophet
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import LSTM, Dense
from sklearn.preprocessing import MinMaxScaler
import warnings
warnings.filterwarnings("ignore")

# ----------------------------
# Helper Functions
# ----------------------------

def detect_congestion(df):
    df['Congestion'] = df['Required_Bandwidth'] > df['Allocated_Bandwidth']
    return df['Congestion'].sum()

def compute_z_scores(series):
    return np.abs((series - series.mean()) / series.std())

def simple_arima_forecast(series, steps=24):
    # Simple exponential smoothing approximation for ARIMA(0,1,0)
    preds = [series.iloc[-1]] * steps
    return np.array(preds)

def simple_sarima_forecast(series, steps=24):
    # Seasonal naive: repeat last week's pattern (hourly assumed)
    if len(series) >= 168:  # 7 days * 24 hours
        seasonal = series[-168:].values
        reps = int(np.ceil(steps / 168))
        return np.tile(seasonal, reps)[:steps]
    else:
        return simple_arima_forecast(series, steps)

def prophet_forecast(df, steps=24):
    m = Prophet(daily_seasonality=True, weekly_seasonality=True)
    df_ts = df[['Timestamp', 'Required_Bandwidth']].rename(columns={'Timestamp': 'ds', 'Required_Bandwidth': 'y'})
    m.fit(df_ts)
    future = m.make_future_dataframe(periods=steps, freq='H')
    forecast = m.predict(future)
    return forecast['yhat'].tail(steps).values

def lstm_forecast(series, steps=24):
    try:
        scaler = MinMaxScaler()
        scaled = scaler.fit_transform(series.values.reshape(-1,1))
        look_back = min(24, len(scaled)//2)
        X, y = [], []
        for i in range(look_back, len(scaled)):
            X.append(scaled[i-look_back:i, 0])
            y.append(scaled[i, 0])
        X, y = np.array(X), np.array(y)
        X = X.reshape((X.shape[0], X.shape[1], 1))

        model = Sequential()
        model.add(LSTM(10, input_shape=(look_back, 1)))
        model.add(Dense(1))
        model.compile(optimizer='adam', loss='mse')
        model.fit(X, y, epochs=10, batch_size=4, verbose=0)

        last_batch = scaled[-look_back:].reshape((1, look_back, 1))
        preds = []
        for _ in range(steps):
            pred = model.predict(last_batch, verbose=0)
            preds.append(pred[0,0])
            last_batch = np.append(last_batch[:,1:,:], pred.reshape(1,1,1), axis=1)
        return scaler.inverse_transform(np.array(preds).reshape(-1,1)).flatten()
    except:
        return simple_arima_forecast(series, steps)

def evaluate_models(actual, pred_dict):
    results = {}
    for name, pred in pred_dict.items():
        if len(pred) != len(actual):
            pred = pred[:len(actual)]
        rmse = np.sqrt(mean_squared_error(actual, pred))
        results[name] = rmse
    return results

def get_congested_hours(df):
    df['Hour'] = pd.to_datetime(df['Timestamp']).dt.hour
    hourly = df.groupby('Hour').agg({
        'Required_Bandwidth': 'mean',
        'Allocated_Bandwidth': 'mean'
    }).reset_index()
    hourly['Shortfall'] = hourly['Required_Bandwidth'] - hourly['Allocated_Bandwidth']
    hourly['Risk'] = pd.cut(hourly['Shortfall'],
                            bins=[-np.inf, 0, 2, 5, np.inf],
                            labels=['None', 'Low', 'Medium', 'High'])
    return hourly[hourly['Shortfall'] > 0].sort_values('Shortfall', ascending=False)

def generate_report_pdf(df, best_model, peak_forecast, peak_time, rec_capacity, congested_table):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Weekly QoS Bandwidth Monitoring Report", ln=True, align='C')
    pdf.ln(10)
    pdf.cell(200, 10, txt=f"Best Model: {best_model}", ln=True)
    pdf.cell(200, 10, txt=f"Peak Forecast: {peak_forecast:.2f} Mbps", ln=True)
    pdf.cell(200, 10, txt=f"Peak Time: {peak_time}", ln=True)
    pdf.cell(200, 10, txt=f"Recommended Capacity: {rec_capacity:.2f} Mbps (+20%)", ln=True)
    pdf.ln(10)
    pdf.cell(200, 10, txt="Congested Hours (Risk Level):", ln=True)
    for _, row in congested_table.iterrows():
        pdf.cell(200, 10, txt=f"Hour {int(row['Hour'])}: Shortfall = {row['Shortfall']:.2f} Mbps ({row['Risk']})", ln=True)
    return pdf

# ----------------------------
# Streamlit App
# ----------------------------

st.set_page_config(page_title="Bandwidth Allocation Dashboard", layout="wide")
st.title("📶 Bandwidth Allocation & Forecasting Dashboard")

# 1) CSV Upload
st.header("1. Upload QoS Dataset")
uploaded_file = st.file_uploader("Drag & drop or click to upload CSV", type=["csv"])

if uploaded_file is not None:
    df = pd.read_csv(uploaded_file)
    required_cols = ['Timestamp', 'Location', 'Signal_Strength', 'Latency', 'Required_Bandwidth', 'Allocated_Bandwidth']
    if not all(col in df.columns for col in required_cols):
        st.error("CSV must contain: Timestamp, Location, Signal_Strength, Latency, Required_Bandwidth, Allocated_Bandwidth")
    else:
        df['Timestamp'] = pd.to_datetime(df['Timestamp'], errors='coerce')
        df = df.dropna(subset=['Timestamp'])
        df = df.sort_values('Timestamp').reset_index(drop=True)

        # KPI Cards
        st.header("2. Key Performance Indicators")
        col1, col2, col3, col4 = st.columns(4)
        total_readings = len(df)
        avg_signal = df['Signal_Strength'].mean()
        avg_latency = df['Latency'].mean()
        congestion_events = detect_congestion(df)
        col1.metric("Total Readings", total_readings)
        col2.metric("Avg Signal (dBm)", f"{avg_signal:.1f}")
        col3.metric("Avg Latency (ms)", f"{avg_latency:.1f}")
        col4.metric("Congestion Events", congestion_events)

        # Prepare data for forecasting
        series = df.set_index('Timestamp')['Required_Bandwidth']
        actual_test = series[-24:]  # last 24 hours as test

        # 2) Model Performance
        st.header("3. Model Performance Comparison")
        with st.spinner("Running forecasting models... (approx. 15-30 sec)"):
            pred_arima = simple_arima_forecast(series, 24)
            pred_sarima = simple_sarima_forecast(series, 24)
            pred_prophet = prophet_forecast(df, 24)
            pred_lstm = lstm_forecast(series, 24)

            model_preds = {
                'ARIMA': pred_arima,
                'SARIMA': pred_sarima,
                'Prophet': pred_prophet,
                'LSTM': pred_lstm
            }
            scores = evaluate_models(actual_test.values, model_preds)
            best_model = min(scores, key=scores.get)
            best_forecast = model_preds[best_model]

        fig, ax = plt.subplots()
        models = list(scores.keys())
        rmse_vals = [scores[m] for m in models]
        bars = ax.bar(models, rmse_vals, color=['skyblue','lightgreen','salmon','gold'])
        ax.set_ylabel('RMSE (Lower = Better)')
        ax.set_title('Model Comparison (RMSE)')
        for bar, val in zip(bars, rmse_vals):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.01, f'{val:.2f}', ha='center')
        st.pyplot(fig)

        # 4) Bandwidth Allocation vs Demand
        st.header("4. Bandwidth Allocation vs Demand")
        fig2, ax2 = plt.subplots(figsize=(10,4))
        ax2.fill_between(df['Timestamp'], df['Required_Bandwidth'], label='Demand', alpha=0.6)
        ax2.plot(df['Timestamp'], df['Allocated_Bandwidth'], label='Allocated', color='red')
        ax2.set_ylabel('Bandwidth (Mbps)')
        ax2.legend()
        st.pyplot(fig2)

        # 5) Forecast vs Actual
        st.header("5. Forecast vs Actual (Last 24h)")
        fig3, ax3 = plt.subplots(figsize=(10,4))
        ax3.plot(actual_test.index, actual_test.values, label='Actual', marker='o')
        ax3.plot(actual_test.index, best_forecast, label=f'Forecast ({best_model})', linestyle='--')
        ax3.set_ylabel('Bandwidth (Mbps)')
        ax3.legend()
        st.pyplot(fig3)

        # 6) Capacity Recommendation
        peak_idx = np.argmax(best_forecast)
        peak_forecast_val = best_forecast[peak_idx]
        peak_time = actual_test.index[peak_idx]
        recommended_capacity = peak_forecast_val * 1.2
        st.header("6. Capacity Recommendation")
        st.write(f"**Peak Forecast**: {peak_forecast_val:.2f} Mbps")
        st.write(f"**Peak Time**: {peak_time.strftime('%Y-%m-%d %H:%M')}")
        st.success(f"✅ **Recommended Capacity**: {recommended_capacity:.2f} Mbps (+20% buffer)")

        # 7) Congested Hours Table
        st.header("7. Congested Hours (Risk Levels)")
        congested_df = get_congested_hours(df)
        st.dataframe(congested_df[['Hour', 'Shortfall', 'Risk']])

        # 8) What-If Scenario
        st.header("8. What-If Capacity Simulation")
        current_alloc = df['Allocated_Bandwidth'].mean()
        new_cap = st.slider("Adjust Allocated Bandwidth (Mbps)", 0.0, 20.0, float(current_alloc))
        df_sim = df.copy()
        df_sim['Allocated_Bandwidth'] = new_cap
        new_congestion = detect_congestion(df_sim)
        st.metric("Simulated Congestion Events", new_congestion, delta=new_congestion - congestion_events)

        # 9) Anomaly Detection
        st.header("9. Anomaly Detection (Demand Spikes)")
        df['Z_Score'] = compute_z_scores(df['Required_Bandwidth'])
        anomalies = df[df['Z_Score'] > 2.5].copy()
        anomalies = anomalies[['Timestamp', 'Required_Bandwidth', 'Z_Score']].sort_values('Z_Score', ascending=False)
        st.dataframe(anomalies)

        # 10) Report Download
        st.header("10. Download Report")
        pdf = generate_report_pdf(df, best_model, peak_forecast_val, peak_time, recommended_capacity, congested_df)
        pdf_output = pdf.output(dest='S').encode('latin1')

        # Excel
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Raw Data', index=False)
            congested_df.to_excel(writer, sheet_name='Congested Hours', index=False)
            anomalies.to_excel(writer, sheet_name='Anomalies', index=False)
        excel_data = output.getvalue()

        col_a, col_b = st.columns(2)
        col_a.download_button("📥 Download PDF Report", data=pdf_output, file_name="QoS_Report.pdf", mime="application/pdf")
        col_b.download_button("📊 Download Excel Report", data=excel_data, file_name="QoS_Report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    st.info("👆 Please upload a CSV file to begin.")